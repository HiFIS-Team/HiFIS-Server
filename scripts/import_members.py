"""기존 SaaS(브로제이·다짐) → HiFIS 회원 일괄 import 스크립트.

용도:
    docker compose exec app python scripts/import_members.py <엑셀파일> [--dry-run]

엑셀 표준 컬럼 (1행 헤더, 사장님 데이터 받고 매핑 조정 필요할 수 있음):
    지점명 | 이름 | 전화번호 | 성별 | 생년월일 | 주소 | 회원권 | 시작일 |
    만기일 | 결제방법 | 결제금액 | 가입일 | 마케팅동의

    - 지점명: "피트니스스타 화순점" 등 (Branch.name 정확히 일치)
    - 성별: M / F / 남 / 여
    - 생년월일: 1990-01-01 또는 datetime
    - 회원권: "1개월" 등 (해당 지점 MembershipPass.name 일치)
    - 결제방법: CASH / CARD / TRANSFER / GIFT_CARD
    - 마케팅동의: TRUE / FALSE / Y / N / O / X / (빈 칸=False)
    - 가입일: 옛 SaaS의 원본 가입일 (created_at 그대로 박음 - D+N 트리거 자동 차단용)

동작:
    1. 엑셀 파싱
    2. 지점·회원권 lookup (이름 → UUID)
    3. 검증 (필수 컬럼·형식·매핑)
    4. INSERT (LMS·Push 발송 0)
    5. 실패 row는 import_failed.csv로 저장

옵션:
    --dry-run: 검증만 하고 INSERT 안 함 (사장님 데이터 사전 확인용)
"""
import argparse
import csv
import sys
from datetime import date, datetime
from pathlib import Path
from zoneinfo import ZoneInfo

# 모든 모델 로드 (FK 해석)
import app.main  # noqa: F401

from openpyxl import load_workbook

from app.db.session import SessionLocal
from app.models.branch import Branch
from app.models.passes.membership import MembershipPass
from app.services.registrations.member import bulk_import_members_silent
from app.utils.validators import is_valid_phone, normalize_phone

_KST = ZoneInfo("Asia/Seoul")

# 엑셀 헤더 → 내부 키 (사장님 엑셀 헤더 다르면 여기 매핑만 수정)
_COLUMN_MAP = {
    "지점명": "branch_name",
    "이름": "name",
    "전화번호": "phone",
    "성별": "gender",
    "생년월일": "birth_date",
    "주소": "address",
    "회원권": "pass_name",
    "시작일": "start_date",
    "만기일": "end_date",
    "결제방법": "payment_method",
    "결제금액": "final_price",
    "가입일": "created_at",
    "마케팅동의": "agreed_marketing",
}

_GENDER_MAP = {"M": "M", "F": "F", "남": "M", "여": "F", "남자": "M", "여자": "F"}
_PAYMENT_MAP = {
    "CASH": "CASH", "현금": "CASH",
    "CARD": "CARD", "카드": "CARD",
    "TRANSFER": "TRANSFER", "계좌이체": "TRANSFER", "이체": "TRANSFER",
    "GIFT_CARD": "GIFT_CARD", "상품권": "GIFT_CARD",
}
_BOOL_MAP = {
    "TRUE": True, "FALSE": False,
    "Y": True, "N": False, "O": True, "X": False,
    "동의": True, "거부": False, "예": True, "아니오": False,
    "1": True, "0": False,
}


def _to_date(val) -> date:
    """엑셀 셀 → date 변환. datetime/date 그대로, 문자열은 파싱."""
    if val is None or val == "":
        raise ValueError("날짜 비어있음")
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
            try:
                return datetime.strptime(val.strip(), fmt).date()
            except ValueError:
                continue
    raise ValueError(f"날짜 형식 알 수 없음: {val!r}")


def _to_datetime_kst(val) -> datetime:
    """엑셀 셀 → 가입일 (KST timestamptz)"""
    d = _to_date(val)
    return datetime(d.year, d.month, d.day, 12, 0, 0, tzinfo=_KST)


def _to_bool(val) -> bool:
    """엑셀 셀 → bool. 빈 값은 False (보수적)"""
    if val is None or val == "":
        return False
    if isinstance(val, bool):
        return val
    s = str(val).strip().upper()
    if s in _BOOL_MAP:
        return _BOOL_MAP[s]
    return False


def _to_int(val) -> int:
    if val is None or val == "":
        return 0
    if isinstance(val, (int, float)):
        return int(val)
    s = str(val).strip().replace(",", "").replace("원", "")
    return int(float(s))


def parse_excel(path: str) -> list[dict]:
    """엑셀 1행을 헤더로 보고 dict 리스트로 반환."""
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    headers = next(rows_iter)
    header_keys = []
    for h in headers:
        if h is None:
            header_keys.append(None)
            continue
        key = _COLUMN_MAP.get(str(h).strip())
        header_keys.append(key)

    raw_rows = []
    for row in rows_iter:
        if all(v is None or v == "" for v in row):
            continue  # 빈 행 스킵
        d = {}
        for key, val in zip(header_keys, row):
            if key is not None:
                d[key] = val
        raw_rows.append(d)
    return raw_rows


def validate_and_map(
    raw_rows: list[dict],
    branches: dict[str, "Branch"],
    passes_by_branch: dict[str, dict[str, "MembershipPass"]],
) -> tuple[list[dict], list[dict]]:
    """엑셀 raw row → Member INSERT 가능한 dict로 검증·변환.

    반환: (검증 통과 row 리스트, 실패 row 리스트)
    """
    today = datetime.now(_KST).date()
    validated: list[dict] = []
    failed: list[dict] = []

    for idx, row in enumerate(raw_rows, start=2):  # 엑셀 2행부터 데이터
        name = row.get("name") or "?"
        try:
            # 지점 매핑
            branch_name = (row.get("branch_name") or "").strip()
            if branch_name not in branches:
                raise ValueError(f"지점 없음: {branch_name!r}")
            branch = branches[branch_name]

            # 회원권 매핑
            pass_name = (row.get("pass_name") or "").strip()
            passes = passes_by_branch.get(branch.id, {})
            if pass_name not in passes:
                raise ValueError(
                    f"회원권 없음: {pass_name!r} ({branch_name})"
                )
            membership_pass = passes[pass_name]

            # 전화번호
            phone_raw = str(row.get("phone") or "").strip()
            if not is_valid_phone(phone_raw):
                raise ValueError(f"전화번호 형식: {phone_raw!r}")
            phone = normalize_phone(phone_raw)

            # 성별
            gender_raw = str(row.get("gender") or "").strip()
            gender = _GENDER_MAP.get(gender_raw.upper()) or _GENDER_MAP.get(gender_raw)
            if not gender:
                raise ValueError(f"성별: {gender_raw!r}")

            # 결제 방법
            pm_raw = str(row.get("payment_method") or "").strip()
            payment_method = _PAYMENT_MAP.get(pm_raw.upper()) or _PAYMENT_MAP.get(pm_raw)
            if not payment_method:
                raise ValueError(f"결제방법: {pm_raw!r}")

            # 날짜들
            birth_date = _to_date(row.get("birth_date"))
            start_date = _to_date(row.get("start_date"))
            end_date = _to_date(row.get("end_date"))
            created_at = _to_datetime_kst(row.get("created_at"))

            if end_date < start_date:
                raise ValueError("end_date < start_date")

            # status: 오늘 기준 활성/만료
            status = "REGISTERED" if end_date >= today else "EXPIRED"

            validated.append({
                "branch_id": branch.id,
                "membership_pass_id": membership_pass.id,
                "name": str(name).strip(),
                "gender": gender,
                "birth_date": birth_date,
                "phone": phone,
                "address": str(row.get("address") or "").strip() or "-",
                "referral": "OTHER",  # 옛 SaaS엔 유입경로 없음 - 기본 OTHER
                "payment_method": payment_method,
                "final_price": _to_int(row.get("final_price")),
                "start_date": start_date,
                "end_date": end_date,
                "motivation": "HEALTH_IMPROVEMENT",  # 옛 SaaS엔 없음 - 기본
                "agreed_terms": True,  # 가입했다는 것 자체가 약관 동의
                "agreed_marketing": _to_bool(row.get("agreed_marketing")),
                "status": status,
                "created_at": created_at,
            })
        except Exception as e:
            failed.append({
                "row": idx,
                "name": name,
                "error": str(e),
            })
    return validated, failed


def save_failed_csv(failed: list[dict], path: str = "import_failed.csv") -> None:
    """실패 row를 CSV로 저장 - 사장님이 보고 엑셀 고치게."""
    if not failed:
        return
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["row", "name", "error"])
        writer.writeheader()
        writer.writerows(failed)
    print(f"[!] 실패 {len(failed)}건 → {path}")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="기존 회원 데이터 일괄 import (LMS·Push 발송 0)"
    )
    parser.add_argument("file", help="엑셀(.xlsx) 파일 경로")
    parser.add_argument(
        "--dry-run", action="store_true",
        help="검증만 수행하고 INSERT 안 함 (사전 확인용)",
    )
    args = parser.parse_args()

    if not Path(args.file).exists():
        print(f"ERROR: 파일 없음: {args.file}", file=sys.stderr)
        sys.exit(1)

    print(f"[1/4] 엑셀 파싱: {args.file}")
    raw_rows = parse_excel(args.file)
    print(f"  → {len(raw_rows)} 행 발견")

    print("[2/4] 지점·회원권 lookup")
    db = SessionLocal()
    try:
        branches = {b.name: b for b in db.query(Branch).all()}
        passes_by_branch: dict = {}
        for p in db.query(MembershipPass).all():
            passes_by_branch.setdefault(p.branch_id, {})[p.name] = p
        print(f"  → 지점 {len(branches)}개, 회원권 총 "
              f"{sum(len(v) for v in passes_by_branch.values())}개")

        print("[3/4] 검증·매핑")
        validated, parse_failed = validate_and_map(
            raw_rows, branches, passes_by_branch,
        )
        print(f"  → 검증 OK: {len(validated)}, 실패: {len(parse_failed)}")

        if parse_failed:
            print("  실패 샘플 (앞 5건):")
            for f in parse_failed[:5]:
                print(f"    row {f['row']}: {f['name']} - {f['error']}")

        if args.dry_run:
            print()
            print("DRY-RUN 종료 - DB 변경 없음")
            save_failed_csv(parse_failed)
            return

        print(f"[4/4] INSERT 시작 ({len(validated)}건)")
        success, insert_failed = bulk_import_members_silent(db, validated)
        print()
        print("=" * 60)
        print(f"✅ 성공: {success}")
        print(f"❌ 실패: {len(parse_failed) + len(insert_failed)} "
              f"(파싱: {len(parse_failed)}, INSERT: {len(insert_failed)})")
        print("=" * 60)

        save_failed_csv(parse_failed + insert_failed)

        if success > 0:
            print()
            print("→ 알림 발송 0건 (silent import)")
            print("→ created_at은 원본 가입일이라 D+N 트리거 옛 회원에겐 자동 미발송")
            print("→ 만기 다가오는 회원은 EXPIRY_SOON 트리거 자연 수신 시작")
    finally:
        db.close()


if __name__ == "__main__":
    main()
