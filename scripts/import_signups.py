"""사장님 신규 등록 회원 일괄 import (HiFIS 신규회원 양식).

브로제이 export(`import_members.py`)와 별개 - 사장님이 export 후 새로 등록한 회원들이
"HiFIS_신규회원 list_<지점>" 형식 .xlsx 로 따로 정리됨.

양식 (한 시트, 좌측 일반회원 + 우측 PT 좌우 분리):
    좌측 (컬럼 0~10): 구분 | 이름 | 성별 | 나이 | 연락처 | 이용기간 | 락카 | 운동복 |
                     최종결제금액 | 방문목적 | 유입경로
    우측 (컬럼 12~22): 구분 | 이름 | 성별 | 나이 | 연락처 | 방문목적 | 유입경로 |
                      세션 | 최종 결제금액 | 상담내용 | 특이사항

좌·우는 서로 다른 회원 (한 행에 일반 1명 + PT 1명 묶여있음).

용도:
    docker compose exec app python scripts/import_signups.py \\
        /tmp/signups.xlsx --branch "피트니스스타 화순점" [--dry-run] [--skip-existing]

옵션:
    --branch    지점명 (Branch.name 정확히 일치)
    --dry-run   검증만, INSERT 안 함
    --skip-existing  이미 DB에 있는 (이름+전화) 매칭 회원은 건너뜀 (중복 방지)

처리:
- 시작일 = 오늘, 만기일 = 이용기간/세션의 N개월에서 자동 계산
- 회원권/락커/운동복/PT 매핑은 공백·대괄호 정규화 후 lookup
- 결제금액은 엑셀 명시값 그대로 파싱 ("385,000원 (VAT)" → 385000)
- 미매핑 회원권명 집계 → 사장님 등록 가이드
"""
import argparse
import csv
import os
import re
import sys
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import NamedTuple
from zoneinfo import ZoneInfo

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import app.main  # noqa: F401

from openpyxl import load_workbook

from app.db.session import SessionLocal
from app.models.branch import Branch
from app.models.passes.clothes import ClothesPass
from app.models.passes.locker import LockerPass
from app.models.passes.membership import MembershipPass
from app.models.passes.pt import PTPass
from app.models.registrations.member import Member
from app.models.registrations.pt_application import PTApplication
from app.utils.validators import is_valid_phone, normalize_phone


_KST = ZoneInfo("Asia/Seoul")

_GENDER_MAP = {"M": "M", "F": "F", "남": "M", "여": "F"}

_REFERRAL_MAP = {
    "네이버": "NAVER",
    "블로그": "BLOG",
    "전단지": "FLYER",
    "인스타": "INSTAGRAM",
    "인스타그램": "INSTAGRAM",
    "현수막": "BANNER",
    "지인소개": "FRIEND",
    "지인추천": "FRIEND",
    "직원소개": "FRIEND",
    "제휴업체": "OTHER",
}

# 방문목적 매핑 (자유 텍스트 → enum). 못 찾으면 NULL.
_MOTIVATION_MAP = {
    "체중감량": "WEIGHT_LOSS",
    "다이어트": "WEIGHT_LOSS",
    "근육증가": "MUSCLE_GAIN",
    "근육 증가": "MUSCLE_GAIN",
    "근력향상": "MUSCLE_GAIN",
    "체력증진": "HEALTH_IMPROVEMENT",
    "건강개선": "HEALTH_IMPROVEMENT",
    "건강 개선": "HEALTH_IMPROVEMENT",
    "운동": "HEALTH_IMPROVEMENT",
    "스트레스해소": "STRESS_RELIEF",
    "외모변화": "APPEARANCE",
    "주변권유": "RECOMMENDATION",
    "부상예방": "INJURY_PREVENTION",
    "체형교정": "POSTURE_CORRECTION",
}


def _normalize_pass(s: str) -> str:
    """이용권 이름 매칭용 정규화:
    - 모든 공백 제거
    - '(' → '[', ')' → ']'  (괄호↔대괄호 차이 무시: (제휴)·[제휴])
    - 'PT' 단어 제거 ('1:1 PT 5회' = '1:1 5회')
    """
    s = "".join(s.split())
    s = s.replace("(", "[").replace(")", "]")
    s = re.sub(r"PT", "", s, flags=re.IGNORECASE)
    return s


def _parse_money(v) -> int | None:
    """'80,000원', '385,000원 (VAT)' → 80000 / 385000"""
    if v is None or v == "":
        return None
    s = str(v)
    digits = re.sub(r"[^\d]", "", s)
    return int(digits) if digits else None


def _extract_months(s: str) -> int | None:
    """이용기간 문자열에서 개월 수 추출. '(제휴) 3개월' → 3, '12개월' → 12"""
    m = re.search(r"(\d+)\s*개월", s)
    if m:
        return int(m.group(1))
    return None


def _add_months(d: date, n: int) -> date:
    """date에 n개월 더하기 (말일 보정 단순화)"""
    y, m = divmod(d.month - 1 + n, 12)
    return date(d.year + y, m + 1, min(d.day, 28))


def _parse_gender(v) -> str | None:
    if not v:
        return None
    s = str(v).strip().upper()
    return _GENDER_MAP.get(s) or _GENDER_MAP.get(s.replace(" ", ""))


def _parse_referral(v) -> tuple[str, str | None]:
    """유입경로 → (enum 값, detail). '기타 (거리우선)' → ('OTHER', '거리우선')"""
    if not v:
        return ("OTHER", None)
    s = str(v).strip()
    # '기타' + 괄호 안 자유 텍스트 패턴
    m = re.match(r"기타\s*\(\s*(.+?)\s*\)", s)
    if m:
        return ("OTHER", m.group(1).strip())
    return (_REFERRAL_MAP.get(s, "OTHER"), None)


def _parse_motivation(v) -> str | None:
    """방문목적 → enum. 못 찾으면 NULL."""
    if not v:
        return None
    s = str(v).strip()
    # 콤마 구분이면 첫 번째만
    first = s.split(",")[0].strip()
    return _MOTIVATION_MAP.get(first)


def _safe_phone(v) -> str | None:
    if not v:
        return None
    s = str(v).strip()
    if is_valid_phone(s):
        return normalize_phone(s)
    return None


def parse_sheet(path: str) -> tuple[list[dict], list[dict]]:
    """한 시트의 좌측 일반회원 + 우측 PT를 dict 리스트로.

    헤더 위치는 첫 행에서 컬럼 이름으로 동적 인식.
    """
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]

    # 좌측 (이용기간 컬럼이 좌측에만 있음) / 우측 (세션 컬럼이 우측에만 있음)으로 영역 구분
    left_h = {h: i for i, h in enumerate(headers) if h and h != "구분"}
    # '구분'·'이름'·'성별'·'나이'·'연락처' 등이 중복으로 양쪽에 있어
    # 첫 번째 인덱스를 좌측, 두 번째를 우측으로 분리
    left_idx = {}
    right_idx = {}
    seen_left = set()
    for i, h in enumerate(headers):
        if not h:
            continue
        if h not in seen_left:
            left_idx[h] = i
            seen_left.add(h)
        else:
            right_idx[h] = i

    members, pts = [], []
    for row in ws.iter_rows(min_row=2, values_only=True):
        # 좌측 일반회원
        ln = row[left_idx["이름"]] if "이름" in left_idx and len(row) > left_idx["이름"] else None
        if ln and str(ln).strip():
            d = {
                "name": str(ln).strip(),
                "gender": _parse_gender(row[left_idx.get("성별", -1)] if "성별" in left_idx else None),
                "phone": _safe_phone(row[left_idx.get("연락처", -1)] if "연락처" in left_idx else None),
                "pass_name": str(row[left_idx["이용기간"]]).strip() if "이용기간" in left_idx and row[left_idx["이용기간"]] else None,
                "locker_name": str(row[left_idx["락카"]]).strip() if "락카" in left_idx and row[left_idx["락카"]] else None,
                "clothes_name": str(row[left_idx["운동복"]]).strip() if "운동복" in left_idx and row[left_idx["운동복"]] else None,
                "final_price": _parse_money(row[left_idx.get("최종결제금액", -1)] if "최종결제금액" in left_idx else None),
                "motivation_raw": row[left_idx.get("방문목적", -1)] if "방문목적" in left_idx else None,
                "referral_raw": row[left_idx.get("유입경로", -1)] if "유입경로" in left_idx else None,
            }
            members.append(d)
        # 우측 PT
        rn = row[right_idx["이름"]] if "이름" in right_idx and len(row) > right_idx["이름"] else None
        if rn and str(rn).strip():
            d = {
                "name": str(rn).strip(),
                "gender": _parse_gender(row[right_idx.get("성별", -1)] if "성별" in right_idx else None),
                "phone": _safe_phone(row[right_idx.get("연락처", -1)] if "연락처" in right_idx else None),
                "pt_name": str(row[right_idx["세션"]]).strip() if "세션" in right_idx and row[right_idx["세션"]] else None,
                "final_price": _parse_money(row[right_idx.get("최종 결제금액", -1)] if "최종 결제금액" in right_idx else None),
                "motivation_raw": row[right_idx.get("방문목적", -1)] if "방문목적" in right_idx else None,
                "referral_raw": row[right_idx.get("유입경로", -1)] if "유입경로" in right_idx else None,
                "notes": str(row[right_idx["특이사항"]]).strip() if "특이사항" in right_idx and row[right_idx["특이사항"]] else None,
            }
            pts.append(d)
    return members, pts


def build_member_row(d: dict, branch, memberships, lockers, clothes, today, warnings: list[str]):
    """좌측 일반회원 dict → Member INSERT용 dict 또는 None."""
    if not d.get("phone"):
        warnings.append(f"{d['name']}: 전화번호 누락 → 제외")
        return None
    if not d.get("pass_name"):
        warnings.append(f"{d['name']}: 이용기간 누락 → 제외")
        return None

    pass_obj = memberships.get((branch.id, _normalize_pass(d["pass_name"])))
    if pass_obj is None:
        warnings.append(f"{d['name']}: 회원권 '{d['pass_name']}' 미매핑 → 제외")
        return None

    # 락커·운동복 - 회원권이 무료제공이면 매핑 스킵 (locker_pass_id=NULL)
    locker_obj = None
    if pass_obj.provides_locker:
        if d.get("locker_name") and d["locker_name"].upper() != "X":
            warnings.append(
                f"{d['name']}: 회원권 '{pass_obj.name}'에 락커 포함 → 락커 매핑 스킵"
            )
    elif d.get("locker_name") and d["locker_name"].upper() != "X":
        # 새 양식은 "3개월"처럼 기간만 → DB의 "락커 3개월" 같은 형태로 lookup
        for candidate in [d["locker_name"], f"락커 {d['locker_name']}"]:
            locker_obj = lockers.get((branch.id, _normalize_pass(candidate)))
            if locker_obj:
                break
        if not locker_obj:
            warnings.append(f"{d['name']}: 락커 '{d['locker_name']}' 미매핑 (NULL로 박음)")

    clothes_obj = None
    if pass_obj.provides_clothes:
        if d.get("clothes_name") and d["clothes_name"].upper() != "X":
            warnings.append(
                f"{d['name']}: 회원권 '{pass_obj.name}'에 운동복 포함 → 운동복 매핑 스킵"
            )
    elif d.get("clothes_name") and d["clothes_name"].upper() != "X":
        for candidate in [d["clothes_name"], f"운동복 {d['clothes_name']}"]:
            clothes_obj = clothes.get((branch.id, _normalize_pass(candidate)))
            if clothes_obj:
                break
        if not clothes_obj:
            warnings.append(f"{d['name']}: 운동복 '{d['clothes_name']}' 미매핑 (NULL로 박음)")

    # 시작·만기
    months = _extract_months(d["pass_name"]) or 1
    start_date = today
    end_date = _add_months(today, months)

    ref, ref_detail = _parse_referral(d.get("referral_raw"))
    return {
        "branch_id": branch.id,
        "membership_pass_id": pass_obj.id,
        "locker_pass_id": locker_obj.id if locker_obj else None,
        "clothes_pass_id": clothes_obj.id if clothes_obj else None,
        "name": d["name"],
        "gender": d.get("gender"),
        "birth_date": None,
        "phone": d["phone"],
        "address": "-",
        "referral": ref,
        "referral_detail": ref_detail,
        "payment_method": None,
        "final_price": d.get("final_price"),
        "total_paid": d.get("final_price"),  # 첫 결제 = 누적
        "start_date": start_date,
        "end_date": end_date,
        "motivation": _parse_motivation(d.get("motivation_raw")),
        "agreed_terms": True,
        "agreed_marketing": False,
        "status": "REGISTERED",
        "created_at": datetime.now(_KST),
    }


def build_pt_row(d: dict, branch, pt_passes, today, warnings: list[str]):
    """우측 PT dict → PTApplication INSERT용 dict 또는 None."""
    if not d.get("phone"):
        warnings.append(f"{d['name']}: 전화번호 누락 → 제외")
        return None
    if not d.get("pt_name"):
        warnings.append(f"{d['name']}: 세션 누락 → 제외")
        return None

    pass_obj = pt_passes.get((branch.id, _normalize_pass(d["pt_name"])))
    if pass_obj is None:
        warnings.append(f"{d['name']}: PT '{d['pt_name']}' 미매핑 → 제외")
        return None

    # PT는 만기 추정 어려움 (회당 진행) → 시작일 = 오늘, 만기일 = +3개월 디폴트
    start_date = today
    end_date = today + timedelta(days=90)

    ref, ref_detail = _parse_referral(d.get("referral_raw"))
    return {
        "branch_id": branch.id,
        "pt_pass_id": pass_obj.id,
        "locker_pass_id": None,
        "clothes_pass_id": None,
        "name": d["name"],
        "gender": d.get("gender"),
        "birth_date": None,
        "phone": d["phone"],
        "address": "-",
        "referral": ref,
        "referral_detail": ref_detail,
        "payment_method": None,
        "final_price": d.get("final_price"),
        "total_paid": d.get("final_price"),  # 첫 결제 = 누적
        "start_date": start_date,
        "end_date": end_date,
        "motivation": _parse_motivation(d.get("motivation_raw")),
        "notes": d.get("notes"),
        "agreed_notice": True,
        "agreed_marketing": False,
        "status": "REGISTERED",
        "created_at": datetime.now(_KST),
    }


def bulk_insert_silent(db, dicts, model):
    success, failed = 0, []
    for idx, data in enumerate(dicts):
        try:
            with db.begin_nested():
                obj = model(**data)
                db.add(obj)
            success += 1
        except Exception as e:
            failed.append({"index": idx, "name": data.get("name", "?"), "error": str(e)})
    db.commit()
    return success, failed


def main():
    parser = argparse.ArgumentParser(
        description="HiFIS 신규회원 list 양식 일괄 import"
    )
    parser.add_argument("file", help="신규회원 xlsx")
    parser.add_argument("--branch", required=True, help="지점명")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument(
        "--skip-existing", action="store_true",
        help="이미 DB에 (이름+전화) 일치 회원이 있으면 건너뜀 (중복 방지)",
    )
    parser.add_argument("--report", default="/tmp/signups_failed.csv")
    args = parser.parse_args()

    if not Path(args.file).exists():
        print(f"ERROR: 파일 없음: {args.file}", file=sys.stderr)
        sys.exit(1)

    print(f"[1/5] 엑셀 파싱: {args.file}")
    excel_members, excel_pts = parse_sheet(args.file)
    print(f"  → 일반회원 {len(excel_members)}, PT {len(excel_pts)}")

    today = datetime.now(_KST).date()

    db = SessionLocal()
    try:
        print("[2/5] 지점·상품 lookup")
        branch = db.query(Branch).filter(Branch.name == args.branch).first()
        if branch is None:
            print(f"ERROR: 지점 없음: {args.branch!r}", file=sys.stderr)
            sys.exit(1)

        memberships = {
            (p.branch_id, _normalize_pass(p.name)): p
            for p in db.query(MembershipPass).filter(MembershipPass.branch_id == branch.id).all()
        }
        pt_passes = {
            (p.branch_id, _normalize_pass(p.name)): p
            for p in db.query(PTPass).filter(PTPass.branch_id == branch.id).all()
        }
        lockers = {
            (p.branch_id, _normalize_pass(p.name)): p
            for p in db.query(LockerPass).filter(LockerPass.branch_id == branch.id).all()
        }
        clothes = {
            (p.branch_id, _normalize_pass(p.name)): p
            for p in db.query(ClothesPass).filter(ClothesPass.branch_id == branch.id).all()
        }
        print(f"  → 회원권 {len(memberships)}, PT {len(pt_passes)}, 락커 {len(lockers)}, 운동복 {len(clothes)}")

        # 중복 검사용 (skip-existing)
        existing_member_keys = set()
        existing_pt_keys = set()
        if args.skip_existing:
            for m in db.query(Member).filter(Member.branch_id == branch.id).all():
                existing_member_keys.add((m.name, m.phone))
            for p in db.query(PTApplication).filter(PTApplication.branch_id == branch.id).all():
                existing_pt_keys.add((p.name, p.phone))

        print("[3/5] 매핑·검증")
        warnings = []
        member_rows, pt_rows = [], []
        skipped_existing = 0

        for d in excel_members:
            if args.skip_existing and (d["name"], d.get("phone")) in existing_member_keys:
                skipped_existing += 1
                continue
            row = build_member_row(d, branch, memberships, lockers, clothes, today, warnings)
            if row:
                member_rows.append(row)

        for d in excel_pts:
            if args.skip_existing and (d["name"], d.get("phone")) in existing_pt_keys:
                skipped_existing += 1
                continue
            row = build_pt_row(d, branch, pt_passes, today, warnings)
            if row:
                pt_rows.append(row)

        print(f"  → Member 검증 OK: {len(member_rows)}")
        print(f"  → PT 검증 OK: {len(pt_rows)}")
        print(f"  → 경고: {len(warnings)}")
        if args.skip_existing:
            print(f"  → 중복 skip: {skipped_existing}")
        if warnings:
            print("  경고 샘플 (앞 15개):")
            for w in warnings[:15]:
                print(f"    {w}")

        if args.dry_run:
            print()
            print("DRY-RUN 종료 - DB 변경 없음")
            return

        print(f"[4/5] Member INSERT ({len(member_rows)}건)")
        m_ok, m_fail = bulk_insert_silent(db, member_rows, Member)
        print(f"  → 성공: {m_ok}, 실패: {len(m_fail)}")

        print(f"[5/5] PT INSERT ({len(pt_rows)}건)")
        p_ok, p_fail = bulk_insert_silent(db, pt_rows, PTApplication)
        print(f"  → 성공: {p_ok}, 실패: {len(p_fail)}")

        print()
        print("=" * 60)
        print(f"✅ Member: {m_ok}, PT: {p_ok}")
        print(f"❌ 실패·경고: {len(warnings) + len(m_fail) + len(p_fail)}")
        if args.skip_existing:
            print(f"⏭️  중복 skip: {skipped_existing}")
        print("=" * 60)
    finally:
        db.close()


if __name__ == "__main__":
    main()
